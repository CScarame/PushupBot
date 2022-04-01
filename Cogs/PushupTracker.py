#############
## PushupTracker.py
## Author: Chris Scaramella
## Date: 1/1/2022
#############
## This cog has all of the main commands for the Pushup Tracking
#############

import discord
from discord.ext import commands
from openpyxl.utils import get_column_letter

import json
import googleHandler
import datetime


def setup(bot):
    bot.add_cog(PushupTracker(bot))

class PushupTracker(commands.Cog):
    def __init__(self, bot):
        self.G = googleHandler.googleHandler("1K9dj5yqQDW7MjXK1J01U-WiVTSJtIns_YHG_s1Shp_U")
        self.bot = bot
        self.PushupsSheetId = 1560699689
        self.SitupsSheetId = 1172496015

    def get_labels(self):
        self.userlist = self.G.read("Pushups!1:1")[0]
        self.datelist = self.G.read("Pushups!A:A")

    def add_user(self, user_id):
        self.G.add_column(self.PushupsSheetId)
        self.G.add_column(self.SitupsSheetId)
        self.userlist.append(f"{user_id}")
        self.G.write(f"Pushups!{get_column_letter(len(self.userlist))}1", 
                    [[f"{user_id}"]])
        self.G.write(f"Situps!{get_column_letter(len(self.userlist))}1", 
                    [[f"{user_id}"]])
        self.G.write(f"Pushups!{get_column_letter(len(self.userlist))}{len(self.datelist)}",
                    [[f'=SUMIFS(RawData!C:C,RawData!B:B,{get_column_letter(len(self.userlist))}' +
                    f'$1,ArrayFormula(TO_DATE(DATEVALUE(RawData!A:A))), "="&TO_DATE($A{len(self.datelist)})' +
                    f',RawData!D:D,"=pushups")']])
        self.G.write(f"Situps!{get_column_letter(len(self.userlist))}{len(self.datelist)}",
                    [[f'=SUMIFS(RawData!C:C,RawData!B:B,{get_column_letter(len(self.userlist))}' +
                    f'$1,ArrayFormula(TO_DATE(DATEVALUE(RawData!A:A))), "="&TO_DATE($A{len(self.datelist)})' +
                    f',RawData!D:D,"=situps")']])
        return

    def add_date(self, curr_date):
        self.datelist.append([curr_date])
        new_line = [curr_date]
        for index, user in enumerate(self.userlist):
            if index != 0:
                new_line.append(f'=SUMIFS(RawData!C:C,RawData!B:B,{get_column_letter(index+1)}$1,' + 
                f'ArrayFormula(TO_DATE(DATEVALUE(RawData!A:A))), "="&TO_DATE($A{len(self.datelist)}) ,RawData!D:D,"=pushups")')
        self.G.append(f"Pushups!A:{get_column_letter(len(self.userlist))}", [new_line])
        for line in new_line:
            line.replace("pushups", "situps")
        self.G.append(f"Situps!A:{get_column_letter(len(self.userlist))}", [new_line])
        return

    @commands.command(help="Check to see how many push ups you have done today.")
    async def daily(self,ctx):
        msg = ""
        self.get_labels()

        ## Check to see if its a new user, and add a column if true
        if f"{ctx.author.id}" not in self.userlist: 
            self.add_user(ctx.author.id)
            msg += "Welcome to Pushup Bot!\n"

        ## Check to see if it is a new day
        curr_date = datetime.datetime.now().strftime("%m/%d/%Y")
        if [curr_date] not in self.datelist:
            self.add_date(curr_date)

        ## Check the total push ups done today
        total = self.G.read(f"Pushups!{get_column_letter(self.userlist.index(f'{ctx.author.id}')+1)}{len(self.datelist)}")
        msg += f"You have done a total of **{total[0][0]}** pushups today."

        await ctx.send(msg)

    @commands.command(help="Track some pushups")
    async def track(self, ctx, *, number):
        msg = ""
        self.get_labels()

        ## Check to see if its a new user, and add a column if true
        if f"{ctx.author.id}" not in self.userlist: 
            self.add_user(ctx.author.id)
            msg += "Welcome to Pushup Bot!\n"

        ## Check to see if it is a new day
        curr_date = datetime.datetime.now().strftime("%m/%d/%Y")
        if [curr_date] not in self.datelist:
            self.add_date(curr_date)
        
        ## Add the new pushups to the data
        self.G.append("RawData!A:C", 
                        [[datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S"), f"{ctx.author.id}", number, "pushups"]])

        ## Check the total push ups done today
        total = self.G.read(f"Pushups!{get_column_letter(self.userlist.index(f'{ctx.author.id}')+1)}{len(self.datelist)}")
        msg += f"Added {number} pushups.\nYou have done a total of **{total[0][0]}** pushups today."
        await ctx.send(msg)

    @commands.command(help="Track pushup shorthand")
    async def p(self, ctx, *, number):
        msg = ""
        self.get_labels()

        ## Check to see if its a new user, and add a column if true
        if f"{ctx.author.id}" not in self.userlist: 
            self.add_user(ctx.author.id)
            msg += "Welcome to Pushup Bot!\n"

        ## Check to see if it is a new day
        curr_date = datetime.datetime.now().strftime("%m/%d/%Y")
        if [curr_date] not in self.datelist:
            self.add_date(curr_date)
        
        ## Add the new pushups to the data
        self.G.append("RawData!A:D", 
                        [[datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S"), f"{ctx.author.id}", number,"pushups"]])

        ## Check the total push ups done today
        total = self.G.read(f"Pushups!{get_column_letter(self.userlist.index(f'{ctx.author.id}')+1)}{len(self.datelist)}")
        msg += f"Added {number} pushups.\nYou have done a total of **{total[0][0]}** pushups today."
        await ctx.send(msg)

    @commands.command(help="Track situp shorthand")
    async def p(self, ctx, *, number):
        msg = ""
        self.get_labels()

        ## Check to see if its a new user, and add a column if true
        if f"{ctx.author.id}" not in self.userlist: 
            self.add_user(ctx.author.id)
            msg += "Welcome to Pushup Bot!\n"

        ## Check to see if it is a new day
        curr_date = datetime.datetime.now().strftime("%m/%d/%Y")
        if [curr_date] not in self.datelist:
            self.add_date(curr_date)
        
        ## Add the new pushups to the data
        self.G.append("RawData!A:D", 
                        [[datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S"), f"{ctx.author.id}", number,"situps"]])

        ## Check the total push ups done today
        total = self.G.read(f"Pushups!{get_column_letter(self.userlist.index(f'{ctx.author.id}')+1)}{len(self.datelist)}")
        msg += f"Added {number} pushups.\nYou have done a total of **{total[0][0]}** pushups today."
        await ctx.send(msg)
    